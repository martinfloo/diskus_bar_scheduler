import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Font, Side
from datetime import datetime, timedelta
import random
from openpyxl.utils import get_column_letter


class BarScheduler:
    def __init__(self):
        self.YEAR = 2024
        self.MONTH = 11  # November
        self.MONTH_NAMES = {
            1: "January",
            2: "February",
            3: "March",
            4: "April",
            5: "May",
            6: "June",
            7: "July",
            8: "August",
            9: "September",
            10: "October",
            11: "November",
            12: "December",
        }
        self.MONTH_NAME = self.MONTH_NAMES[self.MONTH]
        self.USERPATH = "/Users/martin/Desktop/"
        self.FILEPATH = self.USERPATH + f"{self.MONTH_NAME} (Svar) - Skjemasvar 1.csv"
        self.shifts = {
            "opening": {"color": "FFB4C6", "time": "12:30-17:00"},
            "middle": {"color": "B4D7FF", "time": "16:50-20:30"},
            "closing": {"color": "C6FFB4", "time": "20:20-00:30"},
        }
        self.weekend_color = "ECECEC"  # Light gray for weekends
        self.no_reply_color = "404040"  # Dark gray for no replies
        self.manual_review = []
        self.unmatched_availability = {}
        self.no_reply_members = set()

    def is_weekend(self, date_str):
        try:
            day = int(date_str.split(".")[0])
            if day < 1 or day > 31:
                return True
            date = datetime(self.YEAR, self.MONTH, day)
            return date.weekday() >= 5  # Saturday is 5, Sunday is 6
        except:
            return True

    def is_monday(self, date_str):
        try:
            day = int(date_str.split(".")[0])
            if day < 1 or day > 31:  # Using 31 as a safe upper bound
                return False
            date = datetime(self.YEAR, self.MONTH, day)
            return date.weekday() == 0  # Monday is 0 in Python's weekday()
        except:
            return False

    def get_available_shifts(self, date):
        """Helper method to get available shifts for a given date"""
        if self.is_monday(date):
            return ["opening", "middle"]  # No closing shift on Mondays
        return ["opening", "middle", "closing"]

    def get_next_weekend_dates(self, current_date, next_date):
        try:
            current_day = int(current_date.split(".")[0])
            next_day = int(next_date.split(".")[0])
            weekend_dates = []

            current = datetime(self.YEAR, self.MONTH, current_day)
            while current.weekday() < 5 and current.day < next_day:
                current = current + timedelta(days=1)

            if current.day < next_day:
                weekend_dates.append(
                    f"{current.day}. {self.MONTH_NAME[:3].lower()}"
                )  # Changed from "nov"
                if current.day + 1 < next_day:
                    next_day_date = current + timedelta(days=1)
                    if next_day_date.month == self.MONTH:
                        weekend_dates.append(
                            f"{next_day_date.day}. {self.MONTH_NAME[:3].lower()}"
                        )  # Changed from "nov"

            return weekend_dates
        except:
            return []

    def get_staff_requirement(self, date_str, shift_type):
        """Get required staff count based on day and shift"""
        try:
            day = int(date_str.split(".")[0])
            date = datetime(self.YEAR, self.MONTH, day)
            weekday = date.weekday()  # 0 is Monday, 6 is Sunday

            requirements = {
                0: {"opening": 2, "middle": 2, "closing": 0},  # Monday
                1: {"opening": 2, "middle": 3, "closing": 3},  # Tuesday
                2: {"opening": 2, "middle": 3, "closing": 3},  # Wednesday
                3: {"opening": 2, "middle": 2, "closing": 2},  # Thursday
                4: {"opening": 2, "middle": 3, "closing": 3},  # Friday
            }

            return requirements.get(weekday, {}).get(shift_type, 0)
        except:
            return 2  # Default fallback

    def find_member_match(self, input_name, member_list):
        def normalize_name(name):
            return "".join(c.lower() for c in name if c.isalnum())

        def name_similarity(name1, name2):
            n1 = normalize_name(name1)
            n2 = normalize_name(name2)

            if n1 == n2:
                return 1.0

            parts1 = set(normalize_name(p) for p in name1.split())
            parts2 = set(normalize_name(p) for p in name2.split())

            # Strong first name matching regardless of last name
            if parts1 and parts2:
                first1 = normalize_name(name1.split()[0])
                first2 = normalize_name(name2.split()[0])
                if first1 == first2:
                    return 0.95  # High confidence for exact first name match

            # Fall back to overall similarity if first names don't match
            common_parts = parts1.intersection(parts2)
            total_parts = parts1.union(parts2)

            if common_parts:
                return len(common_parts) / len(total_parts)

            return 0.0

        for member in member_list:
            if member.lower() == input_name.lower():
                return member

        best_match = None
        best_score = 0

        for member in member_list:
            score = name_similarity(input_name, member)
            if score > best_score and score >= 0.5:
                best_score = score
                best_match = member

        if best_match:
            if best_score >= 0.8:
                return best_match
            self.manual_review.append(
                {
                    "input_name": input_name,
                    "possible_match": best_match,
                    "confidence": f"{best_score:.2f}",
                }
            )
            return best_match

        self.manual_review.append(
            {
                "input_name": input_name,
                "possible_match": "No match found",
                "confidence": "0.00",
            }
        )
        return input_name

    def parse_shifts(self, cell_value, date_str):
        shifts = []
        if (
            isinstance(cell_value, str)
            and "Kan ikke jobbe denne dagen" not in cell_value
        ):
            if "12:30-17:00" in cell_value:
                shifts.append("opening")
            if "16:50-20:30" in cell_value:
                shifts.append("middle")
            if "20:20-00:30" in cell_value and not self.is_monday(date_str):
                shifts.append("closing")
        # Extra safety check
        if self.is_monday(date_str):
            shifts = [s for s in shifts if s != "closing"]
        return shifts

    def check_consecutive_days(self, schedule, staff_name, current_date, dates):
        current_idx = dates.index(current_date)
        if current_idx > 0:
            prev_date = dates[current_idx - 1]
            if not self.is_weekend(prev_date):
                for shift_type_list in schedule[prev_date].values():
                    if shift_type_list is not None and staff_name in shift_type_list:
                        return True
        return False

    def assign_no_reply_shifts(self, schedule, all_dates, no_reply_members):
        """
        Assigns shifts to members who didn't reply to the availability survey.
        Ensures proper handling of Monday shifts and staff limits.
        """
        for member in no_reply_members:
            shifts_to_assign = 2
            workdays = [date for date in all_dates if not self.is_weekend(date)]
            random.shuffle(workdays)

            for date in workdays:
                if shifts_to_assign <= 0:
                    break

                # Get valid shifts and explicitly exclude closing on Mondays
                valid_shifts = (
                    ["opening", "middle"]
                    if self.is_monday(date)
                    else ["opening", "middle", "closing"]
                )
                random.shuffle(valid_shifts)

                for shift in valid_shifts:
                    # Double-check that we're not assigning closing shifts on Monday
                    if self.is_monday(date) and shift == "closing":
                        continue

                    if schedule[date][shift] is not None and len(
                        schedule[date][shift]
                    ) < self.get_staff_requirement(date, shift):
                        schedule[date][shift].append(member)
                        shifts_to_assign -= 1
                        break

            # Ensure at least one shift is assigned if none were assigned above
            if shifts_to_assign == 2:  # No shifts were assigned
                for date in workdays:
                    valid_shifts = (
                        ["opening", "middle"]
                        if self.is_monday(date)
                        else ["opening", "middle", "closing"]
                    )
                    for shift in valid_shifts:
                        if schedule[date][shift] is not None and len(
                            schedule[date][shift]
                        ) < self.get_staff_requirement(date, shift):
                            schedule[date][shift].append(member)
                            break
                    else:
                        continue
                    break

    def validate_schedule(self, schedule, all_dates):
        """
        Validates and sanitizes the final schedule.
        Ensures no closing shifts on Mondays and maintains shift integrity.
        """
        for date in all_dates:
            # Strictly handle Mondays - no exceptions
            if self.is_monday(date):
                # Force closing to None on Mondays
                schedule[date]["closing"] = None

                # Only validate opening and middle shifts on Mondays
                for shift in ["opening", "middle"]:
                    if schedule[date][shift] is not None:
                        current_staff = schedule[date][shift]
                        required_staff = self.get_staff_requirement(date, shift)
                        if len(current_staff) > required_staff:
                            schedule[date][shift] = current_staff[:required_staff]
            else:
                # Handle non-Monday days
                for shift_type in ["opening", "middle", "closing"]:
                    if schedule[date][shift_type] is not None:
                        current_staff = schedule[date][shift_type]
                        required_staff = self.get_staff_requirement(date, shift_type)
                        if len(current_staff) > required_staff:
                            schedule[date][shift_type] = current_staff[:required_staff]

        # Final safety check - ensure no Monday closing shifts
        for date in all_dates:
            if self.is_monday(date):
                schedule[date]["closing"] = None

        return schedule

    def format_date_str(self, day):
        """Helper method to format date strings consistently"""
        return f"{day}. {self.MONTH_NAME[:3].lower()}"

    def apply_excel_formatting(self, ws, all_dates, all_members):
        """
        Apply formatting while preserving no-reply member styling.
        """
        # Freeze panes
        ws.freeze_panes = "B2"

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Header styling
        header_fill = PatternFill(
            start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
        )
        header_font = Font(bold=True, size=11)

        # Names styling (for those who did reply)
        names_fill = PatternFill(
            start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
        )
        names_font = Font(bold=True, size=11)

        # Style the headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Names column - we'll skip the fill here as it's handled in create_schedule
        for row in range(2, len(all_members) + 2):
            cell = ws.cell(row=row, column=1)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

            # Only apply our formatting if it's not a no-reply member
            name = cell.value
            if name not in self.no_reply_members:
                cell.font = names_font
                cell.fill = names_fill

        # Main grid
        last_col = len(all_dates) + 3
        for row in range(2, len(all_members) + 2):
            for col in range(2, last_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Totals section
        sum_row = len(all_members) + 3
        total_rows = [sum_row, sum_row + 1, sum_row + 2, sum_row + 3]

        for row in total_rows:
            cell = ws.cell(row=row, column=1)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = header_fill

            for col in range(2, last_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        ws.column_dimensions["A"].width = 30
        for col in range(2, last_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Row height
        for row in range(1, len(all_members) + 6):
            ws.row_dimensions[row].height = 22

    def create_schedule(self):
        """
        Creates the complete bar schedule with proper handling of shifts and validations.
        Ensures no closing shifts on Mondays and proper staff distribution.
        """
        # Read member list
        with open(self.USERPATH + "members.txt", "r") as f:
            all_members = [line.strip() for line in f if line.strip()]

        # Read and process CSV
        df = pd.read_csv(self.FILEPATH)
        date_cols = [
            col
            for col in df.columns
            if f"{self.MONTH_NAME[:3].lower()} -" in col.lower()
        ]
        dates = [col.split("[")[-1].split("]")[0].strip() for col in date_cols]

        # Build complete dates list including weekends
        all_dates = []
        current_date = None
        for i, date in enumerate(dates):
            if current_date:
                current_weekday = datetime(
                    self.YEAR, self.MONTH, int(current_date.split(".")[0])
                ).weekday()

                next_weekday = datetime(
                    self.YEAR, self.MONTH, int(date.split(".")[0])
                ).weekday()

                if current_weekday < 5 and next_weekday < 5:
                    all_dates.extend(self.get_next_weekend_dates(current_date, date))
            all_dates.append(date)
            current_date = date

        # Initialize schedule with proper Monday handling
        schedule = {}
        for date in all_dates:
            schedule[date] = {
                "opening": [],
                "middle": [],
                "closing": None if self.is_monday(date) else [],
            }

        staff_availability = {}
        staff_shifts = {member: 0 for member in all_members}

        # Process availability data
        responding_members = set()
        for _, row in df.iterrows():
            input_name = row["Navn og etternavn"]
            matched_name = self.find_member_match(input_name, all_members)
            if matched_name in all_members:
                responding_members.add(matched_name)
                availability = []
                for date, col in zip(dates, date_cols):
                    # Only get valid shifts for the specific day
                    valid_shifts = self.get_available_shifts(date)
                    shifts = [
                        s
                        for s in self.parse_shifts(row[col], date)
                        if s in valid_shifts
                    ]
                    if shifts:
                        availability.append((date, shifts))
                staff_availability[matched_name] = availability

        # Identify no-reply members
        self.no_reply_members = set(all_members) - responding_members

        # First pass: Assign primary shifts
        for staff_name, availability in staff_availability.items():
            if not availability:
                continue

            random.shuffle(availability)
            shifts_assigned = 0

            for date, shifts in availability:
                if shifts_assigned >= 2:
                    break

                if not self.is_weekend(date) and not self.check_consecutive_days(
                    schedule, staff_name, date, all_dates
                ):
                    valid_shifts = self.get_available_shifts(date)
                    random.shuffle(valid_shifts)

                    for shift in valid_shifts:
                        if schedule[date][shift] is not None and len(
                            schedule[date][shift]
                        ) < self.get_staff_requirement(date, shift):
                            schedule[date][shift].append(staff_name)
                            shifts_assigned += 1
                            break

        # Second pass: Fill remaining slots
        for staff_name, availability in staff_availability.items():
            current_shifts = sum(
                1
                for date in all_dates
                for shift_type, staff_list in schedule[date].items()
                if staff_list is not None and staff_name in staff_list
            )

            if current_shifts >= 2:
                continue

            random.shuffle(availability)
            for date, shifts in availability:
                if current_shifts >= 2:
                    break

                if not self.is_weekend(date) and not self.check_consecutive_days(
                    schedule, staff_name, date, all_dates
                ):
                    valid_shifts = self.get_available_shifts(date)
                    random.shuffle(valid_shifts)

                    for shift in valid_shifts:
                        if schedule[date][shift] is not None and len(
                            schedule[date][shift]
                        ) < self.get_staff_requirement(date, shift):
                            schedule[date][shift].append(staff_name)
                            current_shifts += 1
                            break

        # Handle no-reply members
        self.assign_no_reply_shifts(schedule, all_dates, self.no_reply_members)

        # Final validation
        schedule = self.validate_schedule(schedule, all_dates)

        for date in all_dates:
            if self.is_monday(date):
                schedule[date]["closing"] = None  # Force None for Monday closing shifts

        # Create Excel output
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"

        # Add color legend
        legend_row = 1
        ws.cell(row=legend_row, column=len(all_dates) + 5, value="Shift Colors:")
        for idx, (shift_name, info) in enumerate(self.shifts.items(), 1):
            cell = ws.cell(
                row=legend_row + idx,
                column=len(all_dates) + 5,
                value=f"{shift_name.capitalize()} ({info['time']})",
            )
            cell.fill = PatternFill(
                start_color=info["color"], end_color=info["color"], fill_type="solid"
            )

        # Headers
        ws["A1"] = "Name"
        all_dates_with_type = [(date, self.is_weekend(date)) for date in all_dates]
        for idx, (date, is_weekend) in enumerate(all_dates_with_type, 2):
            cell = ws.cell(row=1, column=idx, value=date)
            if is_weekend:
                cell.fill = PatternFill(
                    start_color=self.weekend_color,
                    end_color=self.weekend_color,
                    fill_type="solid",
                )

        # Fill schedule
        for row_idx, name in enumerate(all_members, 2):
            cell = ws.cell(row=row_idx, column=1, value=name)

            if name in self.no_reply_members:
                cell.fill = PatternFill(
                    start_color=self.no_reply_color,
                    end_color=self.no_reply_color,
                    fill_type="solid",
                )
                cell.font = Font(color="FFFFFF")
            elif any(review["possible_match"] == name for review in self.manual_review):
                cell.fill = PatternFill(
                    start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
                )

            for col_idx, (date, is_weekend) in enumerate(all_dates_with_type, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                if is_weekend:
                    cell.fill = PatternFill(
                        start_color=self.weekend_color,
                        end_color=self.weekend_color,
                        fill_type="solid",
                    )
                else:
                    # Add explicit Monday check here
                    if self.is_monday(date):
                        for shift, staff_list in schedule[date].items():
                            if (
                                shift != "closing"
                                and staff_list is not None
                                and name in staff_list
                            ):
                                cell.fill = PatternFill(
                                    start_color=self.shifts[shift]["color"],
                                    end_color=self.shifts[shift]["color"],
                                    fill_type="solid",
                                )
                    else:
                        for shift, staff_list in schedule[date].items():
                            if staff_list is not None and name in staff_list:
                                cell.fill = PatternFill(
                                    start_color=self.shifts[shift]["color"],
                                    end_color=self.shifts[shift]["color"],
                                    fill_type="solid",
                                )

        # Clear any accidental Monday closing shifts
        for date in all_dates:
            if self.is_monday(date):
                col_idx = all_dates.index(date) + 2
                for row_idx in range(2, len(all_members) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.fill.start_color.index == "C6FFB4":  # Closing shift color
                        cell.fill = PatternFill(fill_type=None)

        # Count shifts and add totals
        staff_shifts = {member: 0 for member in all_members}
        for date in dates:
            for shift_type in ["opening", "middle", "closing"]:
                if schedule[date].get(shift_type) is not None:
                    for staff in schedule[date][shift_type]:
                        if staff in staff_shifts:
                            staff_shifts[staff] += 1

        # Add total shifts and availability columns
        ws.cell(row=1, column=len(all_dates) + 2, value="Total Shifts")
        ws.cell(row=1, column=len(all_dates) + 3, value="Available Days")
        for row_idx, name in enumerate(all_members, 2):
            ws.cell(row=row_idx, column=len(all_dates) + 2, value=staff_shifts[name])
            available_count = len(staff_availability.get(name, []))
            ws.cell(row=row_idx, column=len(all_dates) + 3, value=available_count)

        # Create manual review sheet if needed
        if self.manual_review:
            ws_review = wb.create_sheet("Manual Review")
            ws_review["A1"] = "Input Name"
            ws_review["B1"] = "Possible Match"
            ws_review["C1"] = "Confidence"
            ws_review["D1"] = "Available Dates"

            for idx, review in enumerate(self.manual_review, 2):
                ws_review[f"A{idx}"] = review["input_name"]
                ws_review[f"B{idx}"] = review["possible_match"]
                ws_review[f"C{idx}"] = review.get("confidence", "N/A")

                if review["input_name"] in self.unmatched_availability:
                    avail_text = []
                    for date, shifts in self.unmatched_availability[
                        review["input_name"]
                    ].items():
                        if shifts:
                            avail_text.append(f"{date}: {', '.join(shifts)}")
                    ws_review[f"D{idx}"] = "\n".join(avail_text)

        # Create manual review sheet if needed
        if self.manual_review:
            ws_review = wb.create_sheet("Manual Review")
            ws_review["A1"] = "Input Name"
            ws_review["B1"] = "Possible Match"
            ws_review["C1"] = "Confidence"
            ws_review["D1"] = "Available Dates"

            for idx, review in enumerate(self.manual_review, 2):
                ws_review[f"A{idx}"] = review["input_name"]
                ws_review[f"B{idx}"] = review["possible_match"]
                ws_review[f"C{idx}"] = review.get("confidence", "N/A")

                if review["input_name"] in self.unmatched_availability:
                    avail_text = []
                    for date, shifts in self.unmatched_availability[
                        review["input_name"]
                    ].items():
                        if shifts:
                            avail_text.append(f"{date}: {', '.join(shifts)}")
                    ws_review[f"D{idx}"] = "\n".join(avail_text)

        # Add shift counts per day
        sum_row = len(all_members) + 3  # Leave a blank row after the last member
        ws.cell(row=sum_row, column=1, value="SUM OF SHIFTS")

        for col_idx, (date, is_weekend) in enumerate(all_dates_with_type, 2):
            if not is_weekend:
                opening_count = (
                    len(schedule[date]["opening"])
                    if schedule[date]["opening"] is not None
                    else 0
                )
                middle_count = (
                    len(schedule[date]["middle"])
                    if schedule[date]["middle"] is not None
                    else 0
                )
                closing_count = (
                    len(schedule[date]["closing"])
                    if schedule[date]["closing"] is not None
                    else 0
                )

                # Sum of all shifts
                ws.cell(
                    row=sum_row,
                    column=col_idx,
                    value=opening_count + middle_count + closing_count,
                )

                # Individual shift counts
                ws.cell(row=sum_row + 1, column=1, value="OPENING")
                ws.cell(row=sum_row + 1, column=col_idx, value=opening_count)

                ws.cell(row=sum_row + 2, column=1, value="MIDDAY")
                ws.cell(row=sum_row + 2, column=col_idx, value=middle_count)

                ws.cell(row=sum_row + 3, column=1, value="CLOSING")
                ws.cell(row=sum_row + 3, column=col_idx, value=closing_count)

        self.apply_excel_formatting(ws, all_dates, all_members)

        # Save the workbook
        save_path = (
            f"{self.USERPATH}{self.MONTH_NAME.lower()}_schedule_{self.YEAR}.xlsx"
        )
        print(f"Saving schedule to: {save_path}")
        wb.save(save_path)
        # wb.save(f"{self.MONTH_NAME.lower()}_schedule_{self.YEAR}.xlsx")


def main():
    scheduler = BarScheduler()
    scheduler.create_schedule()


if __name__ == "__main__":
    main()

