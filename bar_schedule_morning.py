import random
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MOCK_DATA = False
# CHOOSE YEAR HERE
MONTH = 12  # December


class BarScheduler:
    SHIFT_CONFIG = {
        "morning": {"color": "FFE4B5", "time": "08:45-12:30", "default_staff": 2},
        "opening": {"color": "FFB4C6", "time": "12:30-17:00", "default_staff": 2},
        "middle": {"color": "B4D7FF", "time": "16:50-20:30", "default_staff": 3},
        "closing": {"color": "C6FFB4", "time": "20:20-00:30", "default_staff": 3},
    }

    EXCEL_STYLES = {
        "thin_border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
        "header": {
            "fill": PatternFill(
                start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
            ),
            "font": Font(bold=True, size=11),
        },
        "names": {
            "fill": PatternFill(
                start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
            ),
            "font": Font(bold=True, size=11),
        },
    }
    MONTH_NAMES = {
        1: "January",
        2: "February",
        3: "March",
        4: "April",
        5: "May",
        6: "June",
        7: "July",
        8: "August",
        9: "September",
        10: "October",
        11: "November",
        12: "Desember",
    }

    MIN_CONFIDENCE_THRESHOLD = 0.8
    PARTIAL_MATCH_THRESHOLD = 0.5

    def __init__(self):
        self.YEAR = 2024
        self.MONTH = MONTH
        self.MONTH_NAME = self.MONTH_NAMES[self.MONTH]

        self.USERPATH = "/Users/martin/Desktop/"
        self.FILEPATH = self.USERPATH + f"{self.MONTH_NAME}<"

        if MOCK_DATA:
            self.USERPATH = "/Users/martin/Desktop/bar-scheduler/"
            self.FILEPATH = self.USERPATH + "mock_data.csv"

        base_requirements = {
            0: {"opening": 2, "middle": 2},
            1: {s: c["default_staff"] for s, c in self.SHIFT_CONFIG.items()},
            2: {s: c["default_staff"] for s, c in self.SHIFT_CONFIG.items()},
            3: {"opening": 2, "middle": 2, "closing": 2},
            4: {s: c["default_staff"] for s, c in self.SHIFT_CONFIG.items()},
        }
        self.WEEKDAY_REQUIREMENTS = base_requirements.copy()
        self.weekend_color = "808080"
        self.no_reply_color = "404040"
        self.manual_review = []
        self.unmatched_availability = {}
        self.no_reply_members = set()
        self.morning_shift_dates = set()

    def format_date(self, day):
        return f"{day}. {self.MONTH_NAME[:3].lower()}"

    def get_weekday(self, date_str):
        try:
            day = int(date_str.split(".")[0])
            return datetime(self.YEAR, self.MONTH, day).weekday()
        except ValueError:
            return -1

    def is_weekend(self, date_str):
        return self.get_weekday(date_str) >= 5

    def is_monday(self, date_str):
        return self.get_weekday(date_str) == 0

    def get_available_shifts(self, date):
        valid_shifts = []
        date_str = date.strip()
        if " - " in date_str:
            date_str = date_str.split(" - ")[0]

        if self._is_morning_shift_date(date_str):
            valid_shifts.append("morning")

        if self.is_monday(date):
            valid_shifts.extend(["opening", "middle"])
        elif not self.is_weekend(date):
            valid_shifts.extend(["opening", "middle", "closing"])

        return valid_shifts

    def get_next_weekend_dates(self, current_date, next_date):
        try:
            current_day = int(current_date.split(".")[0])
            next_day = int(next_date.split(".")[0])
            weekend_dates = []

            current = datetime(self.YEAR, self.MONTH, current_day)
            while current.weekday() < 5 and current.day < next_day:
                current = current + timedelta(days=1)

            if current.day < next_day:
                weekend_dates.append("WEEKEND")

            return weekend_dates
        except ValueError:
            return []

    def get_staff_requirement(self, date_str, shift_type):
        weekday = self.get_weekday(date_str)
        return self.WEEKDAY_REQUIREMENTS.get(weekday, {}).get(
            shift_type, self.SHIFT_CONFIG[shift_type]["default_staff"]
        )

    def get_morning_shift_dates(self, df):
        morning_cols = [col for col in df.columns if "kan du ha morgenvakt?" in col]
        morning_dates = set()

        for col in morning_cols:
            try:
                date_part = col.split("[")[-1].split("]")[0]
                morning_dates.add(date_part)
            except IndexError:
                print(f"Warning: Could not parse date from column: {col}")

        return morning_dates

    def find_member_match(self, input_name, member_list):
        def normalize_name(name):
            return "".join(c.lower() for c in name if c.isalnum())

        if input_name.lower() in (m.lower() for m in member_list):
            return next(m for m in member_list if m.lower() == input_name.lower())

        matches = []
        input_normalized = normalize_name(input_name)
        input_parts = set(normalize_name(p) for p in input_name.split())

        for member in member_list:
            member_normalized = normalize_name(member)
            member_parts = set(normalize_name(p) for p in member.split())

            score = 0
            if input_normalized == member_normalized:
                score = 1.0
            elif input_parts and member_parts:
                if normalize_name(input_name.split()[0]) == normalize_name(
                    member.split()[0]
                ):
                    score = 0.95
                else:
                    common_parts = input_parts.intersection(member_parts)
                    score = (
                        len(common_parts) / len(input_parts.union(member_parts))
                        if common_parts
                        else 0
                    )

            if score > 0:
                matches.append((member, score))

        if matches:
            best_match = max(matches, key=lambda x: x[1])
            if best_match[1] >= self.MIN_CONFIDENCE_THRESHOLD:
                return best_match[0]
            elif best_match[1] >= self.PARTIAL_MATCH_THRESHOLD:
                self.manual_review.append(
                    {
                        "input_name": input_name,
                        "possible_match": best_match[0],
                        "confidence": f"{best_match[1]:.2f}",
                    }
                )
                return best_match[0]

        self.manual_review.append(
            {
                "input_name": input_name,
                "possible_match": "No match found",
                "confidence": "0.00",
            }
        )
        return input_name

    def parse_shifts(self, row, date_str):
        shifts = []
        date_check = date_str.split(" - ")[0] if " - " in date_str else date_str

        date_col = next((col for col in row.index if date_str in col), None)
        if date_col:
            cell_value = row[date_col]
            if (
                isinstance(cell_value, str)
                and "Kan ikke jobbe denne dagen" not in cell_value
            ):
                shift_times = {
                    "opening": "12:30-17:00",
                    "middle": "16:50-20:30",
                    "closing": "20:20-00:30",
                }
                for shift_type, time in shift_times.items():
                    if time in cell_value:
                        if not (self.is_monday(date_str) and shift_type == "closing"):
                            shifts.append(shift_type)

        day = int(date_check.split(".")[0])
        month_abbrev = self.MONTH_NAME[:3].lower()
        morning_cols = [col for col in row.index if "kan du ha morgenvakt?" in col]

        for morning_col in morning_cols:
            if f"[{day}. {month_abbrev}]" in morning_col and row[morning_col] == "Ja":
                shifts.append("morning")
                break

        return shifts

    def update_shift_requirements(self, df):
        morning_dates = self.get_morning_shift_dates(df)
        for date in morning_dates:
            try:
                weekday = self.get_weekday(date)
                if weekday in self.WEEKDAY_REQUIREMENTS:
                    current_reqs = self.WEEKDAY_REQUIREMENTS[weekday]
                    current_reqs["morning"] = self.SHIFT_CONFIG["morning"][
                        "default_staff"
                    ]
                    self.WEEKDAY_REQUIREMENTS[weekday] = current_reqs
            except ValueError:
                print(f"Warning: Could not process morning shift date: {date}")

    def check_consecutive_days(self, schedule, staff_name, current_date, dates):
        current_idx = dates.index(current_date)
        if current_idx > 0:
            prev_date = dates[current_idx - 1]
            if not self.is_weekend(prev_date):
                for shift_type_list in schedule[prev_date].values():
                    if shift_type_list is not None and staff_name in shift_type_list:
                        return True
        return False

    def assign_no_reply_shifts(self, schedule, all_dates, no_reply_members):
        workdays = [date for date in all_dates if not self.is_weekend(date)]

        for member in no_reply_members:
            shifts_needed = 2
            random.shuffle(workdays)

            for date in workdays:
                if shifts_needed <= 0:
                    break

                valid_shifts = (
                    ["opening", "middle"]
                    if self.is_monday(date)
                    else ["opening", "middle", "closing"]
                )
                random.shuffle(valid_shifts)

                for shift in valid_shifts:
                    if schedule[date][shift] is not None and len(
                        schedule[date][shift]
                    ) < self.get_staff_requirement(date, shift):
                        schedule[date][shift].append(member)
                        shifts_needed -= 1
                        break

            if shifts_needed == 2:
                for date in workdays:
                    valid_shifts = (
                        ["opening", "middle"]
                        if self.is_monday(date)
                        else ["opening", "middle", "closing"]
                    )
                    for shift in valid_shifts:
                        if schedule[date][shift] is not None and len(
                            schedule[date][shift]
                        ) < self.get_staff_requirement(date, shift):
                            schedule[date][shift].append(member)
                            break
                    else:
                        continue
                    break

    def validate_schedule(self, schedule, all_dates):
        for date in all_dates:
            if self.is_weekend(date):
                continue

            if self.is_monday(date):
                schedule[date]["closing"] = None

            for shift_type, staff_list in schedule[date].items():
                if staff_list is not None:
                    for staff_name in staff_list[:]:
                        if staff_name not in self.no_reply_members:
                            staff_shifts = self.staff_availability.get(staff_name, [])
                            available_shifts = [
                                s
                                for d, shifts in staff_shifts
                                if d == date
                                for s in shifts
                            ]
                            if shift_type not in available_shifts:
                                print(
                                    f"Warning: Removing {staff_name} from {shift_type} on {date}"
                                )
                                staff_list.remove(staff_name)

                    required = self.get_staff_requirement(date, shift_type)
                    if len(staff_list) > required:
                        print(
                            f"Warning: {date} {shift_type} has {len(staff_list)} people, limiting to {required}"
                        )
                        schedule[date][shift_type] = staff_list[:required]

        return schedule

    def apply_excel_formatting(self, ws, all_dates, all_members):
        ws.freeze_panes = "B2"

        for cell in ws[1]:
            cell.fill = self.EXCEL_STYLES["header"]["fill"]
            cell.font = self.EXCEL_STYLES["header"]["font"]
            cell.border = self.EXCEL_STYLES["thin_border"]
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in range(2, len(all_members) + 2):
            cell = ws.cell(row=row, column=1)
            cell.border = self.EXCEL_STYLES["thin_border"]
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

            name = cell.value
            if name not in self.no_reply_members:
                cell.font = self.EXCEL_STYLES["names"]["font"]
                cell.fill = self.EXCEL_STYLES["names"]["fill"]

        last_col = len(all_dates) + 3
        for row in range(2, len(all_members) + 2):
            for col in range(2, last_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.EXCEL_STYLES["thin_border"]
                cell.alignment = Alignment(horizontal="center", vertical="center")

        sum_row = len(all_members) + 3
        total_rows = [sum_row, sum_row + 1, sum_row + 2, sum_row + 3, sum_row + 4]

        for row in total_rows:
            cell = ws.cell(row=row, column=1)
            cell.font = Font(bold=True)
            cell.border = self.EXCEL_STYLES["thin_border"]
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = self.EXCEL_STYLES["header"]["fill"]

            for col in range(2, last_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.EXCEL_STYLES["thin_border"]
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 30
        for col in range(2, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "WEEKEND":
                ws.column_dimensions[get_column_letter(col)].width = 15
            else:
                ws.column_dimensions[get_column_letter(col)].width = 12

        for row in range(1, len(all_members) + 6):
            ws.row_dimensions[row].height = 22

    def assign_shifts(
        self, schedule, dates, staff_availability, all_members, shifts_needed=1
    ):
        for staff_name, availability in staff_availability.items():
            if self._count_shifts(schedule, staff_name) >= shifts_needed:
                continue

            random.shuffle(availability)
            for date, available_shifts in availability:
                if not (
                    self.is_weekend(date)
                    or self.check_consecutive_days(schedule, staff_name, date, dates)
                ):
                    valid_shifts = [
                        shift
                        for shift in available_shifts
                        if shift in self.get_available_shifts(date)
                    ]
                    if self._try_assign_shift(schedule, date, valid_shifts, staff_name):
                        if self._count_shifts(schedule, staff_name) >= shifts_needed:
                            break

        workdays = [d for d in dates if not self.is_weekend(d)]
        for member in self.no_reply_members:
            if self._count_shifts(schedule, member) >= shifts_needed:
                continue

            random.shuffle(workdays)
            for date in workdays:
                valid_shifts = (
                    ["opening", "middle"]
                    if self.is_monday(date)
                    else ["opening", "middle", "closing"]
                )
                if self._try_assign_shift(schedule, date, valid_shifts, member):
                    if self._count_shifts(schedule, member) >= shifts_needed:
                        break

        return schedule

    def _try_assign_shift(self, schedule, date, valid_shifts, staff_name):
        if self._has_shift_on_date(schedule, date, staff_name):
            return False

        if staff_name not in self.no_reply_members:
            staff_shifts = self.staff_availability.get(staff_name, [])
            available_shifts = [
                s for d, shifts in staff_shifts if d == date for s in shifts
            ]
            valid_shifts = [s for s in valid_shifts if s in available_shifts]
            if not valid_shifts:
                return False

        available_shifts = set(valid_shifts)

        if "morning" in available_shifts and schedule[date].get("morning") is not None:
            required = self.get_staff_requirement(date, "morning")
            current = len(schedule[date]["morning"])
            if current < required:
                schedule[date]["morning"].append(staff_name)
                print(f"Assigned {staff_name} to morning shift on {date}")
                return True

        other_shifts = [s for s in valid_shifts if s != "morning"]
        random.shuffle(other_shifts)
        for shift in other_shifts:
            if schedule[date].get(shift) is not None:
                required = self.get_staff_requirement(date, shift)
                current = len(schedule[date][shift])
                if current < required:
                    schedule[date][shift].append(staff_name)
                    print(f"Assigned {staff_name} to {shift} shift on {date}")
                    return True

        return False

    def _count_shifts(self, schedule, staff_name):
        return sum(
            1
            for date in schedule
            for shift_type, staff_list in schedule[date].items()
            if staff_list is not None and staff_name in staff_list
        )

    def _is_morning_shift_date(self, date):
        if " - " in date:
            date = date.split(" - ")[0]
        return date in self.morning_shift_dates

    def initialize_schedule(self, date, morning_shift_dates):
        schedule = {
            "morning": [] if self._is_morning_shift_date(date.strip()) else None,
            "opening": [],
            "middle": [],
            "closing": None if self.is_monday(date) else [],
        }
        return schedule

    def _has_shift_on_date(self, schedule, date, staff_name):
        for shift_type, staff_list in schedule[date].items():
            if staff_list is not None and staff_name in staff_list:
                return True
        return False

    def create_schedule(self):
        try:
            with open(self.USERPATH + "members.txt", "r") as f:
                all_members = [line.strip() for line in f if line.strip()]
        except FileNotFoundError:
            raise FileNotFoundError("Could not find members.txt file")

        df = pd.read_csv(self.FILEPATH)
        self.morning_shift_dates = self.get_morning_shift_dates(df)
        self.update_shift_requirements(df)

        date_cols = [
            col
            for col in df.columns
            if f"{self.MONTH_NAME[:3].lower()} -" in col.lower()
        ]
        work_dates = [col.split("[")[-1].split("]")[0].strip() for col in date_cols]

        all_dates = []
        current_date = None
        for i, date in enumerate(work_dates):
            if current_date:
                current_day = int(current_date.split(".")[0])
                next_day = int(date.split(".")[0])
                current = datetime(self.YEAR, self.MONTH, current_day)
                while current.weekday() < 5 and current.day < next_day:
                    current = current + timedelta(days=1)
                    if current.day < next_day:
                        weekend_date = f"{current.day}. {self.MONTH_NAME[:3].lower()}"
                        all_dates.append(weekend_date)
            all_dates.append(date)
            current_date = date

        schedule = {}
        for date in all_dates:
            if self.is_weekend(date):
                schedule[date] = {shift_type: None for shift_type in self.SHIFT_CONFIG}
            else:
                schedule[date] = self.initialize_schedule(
                    date, self.morning_shift_dates
                )

        staff_availability = {}
        responding_members = set()

        for _, row in df.iterrows():
            input_name = row["Navn og etternavn"]
            matched_name = self.find_member_match(input_name, all_members)
            if matched_name in all_members:
                responding_members.add(matched_name)
                availability = []

                for date in work_dates:
                    shifts = self.parse_shifts(row, date)
                    if shifts:
                        availability.append((date, shifts))
                staff_availability[matched_name] = availability

        self.staff_availability = staff_availability

        self.no_reply_members = set(all_members) - responding_members

        self.assign_shifts(
            schedule, work_dates, staff_availability, all_members, shifts_needed=1
        )
        self.assign_shifts(
            schedule, work_dates, staff_availability, all_members, shifts_needed=2
        )
        schedule = self.validate_schedule(schedule, all_dates)

        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"

        legend_row = 1
        ws.cell(row=legend_row, column=len(all_dates) + 5, value="Shift Colors:")
        for idx, (shift_name, info) in enumerate(self.SHIFT_CONFIG.items(), 1):
            cell = ws.cell(
                row=legend_row + idx,
                column=len(all_dates) + 5,
                value=f"{shift_name.capitalize()} ({info['time']})",
            )
            cell.fill = PatternFill(
                start_color=info["color"], end_color=info["color"], fill_type="solid"
            )

        ws["A1"] = "Name"
        for idx, date in enumerate(all_dates, 2):
            cell = ws.cell(
                row=1, column=idx, value="WEEKEND" if self.is_weekend(date) else date
            )
            if self.is_weekend(date):
                cell.fill = PatternFill(
                    start_color=self.weekend_color,
                    end_color=self.weekend_color,
                    fill_type="solid",
                )

        for row_idx, name in enumerate(all_members, 2):
            cell = ws.cell(row=row_idx, column=1, value=name)

            if name in self.no_reply_members:
                cell.fill = PatternFill(
                    start_color=self.no_reply_color,
                    end_color=self.no_reply_color,
                    fill_type="solid",
                )
                cell.font = Font(color="FFFFFF")
            elif any(review["possible_match"] == name for review in self.manual_review):
                cell.fill = PatternFill(
                    start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
                )

            for col_idx, date in enumerate(all_dates, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                if self.is_weekend(date):
                    cell.fill = PatternFill(
                        start_color=self.weekend_color,
                        end_color=self.weekend_color,
                        fill_type="solid",
                    )
                else:
                    for shift_type in ["morning", "opening", "middle", "closing"]:
                        if (
                            schedule[date].get(shift_type) is not None
                            and name in schedule[date][shift_type]
                        ):
                            cell.fill = PatternFill(
                                start_color=self.SHIFT_CONFIG[shift_type]["color"],
                                end_color=self.SHIFT_CONFIG[shift_type]["color"],
                                fill_type="solid",
                            )

        for date in all_dates:
            if self.is_monday(date):
                col_idx = all_dates.index(date) + 2
                for row_idx in range(2, len(all_members) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.fill.start_color.index == "C6FFB4":
                        cell.fill = PatternFill(fill_type=None)

        staff_shifts = {member: 0 for member in all_members}
        for date in all_dates:
            for shift_type in [
                "morning",
                "opening",
                "middle",
                "closing",
            ]:
                if schedule[date].get(shift_type) is not None:
                    for staff in schedule[date][shift_type]:
                        if staff in staff_shifts:
                            staff_shifts[staff] += 1

        ws.cell(row=1, column=len(all_dates) + 2, value="Total Shifts")
        ws.cell(row=1, column=len(all_dates) + 3, value="Available Days")
        for row_idx, name in enumerate(all_members, 2):
            ws.cell(row=row_idx, column=len(all_dates) + 2, value=staff_shifts[name])
            available_count = len(staff_availability.get(name, []))
            ws.cell(row=row_idx, column=len(all_dates) + 3, value=available_count)

        if self.manual_review:
            ws_review = wb.create_sheet("Manual Review")
            ws_review["A1"] = "Input Name"
            ws_review["B1"] = "Possible Match"
            ws_review["C1"] = "Confidence"
            ws_review["D1"] = "Available Dates"

            for idx, review in enumerate(self.manual_review, 2):
                ws_review[f"A{idx}"] = review["input_name"]
                ws_review[f"B{idx}"] = review["possible_match"]
                ws_review[f"C{idx}"] = review.get("confidence", "N/A")

                if review["input_name"] in self.unmatched_availability:
                    avail_text = []
                    for date, shifts in self.unmatched_availability[
                        review["input_name"]
                    ].items():
                        if shifts:
                            avail_text.append(f"{date}: {', '.join(shifts)}")
                    ws_review[f"D{idx}"] = "\n".join(avail_text)

        sum_row = len(all_members) + 3
        ws.cell(row=sum_row, column=1, value="SUM OF SHIFTS")

        for col_idx, date in enumerate(all_dates, 2):
            if not self.is_weekend(date):
                morning_count = (
                    len(schedule[date]["morning"])
                    if schedule[date]["morning"] is not None
                    else 0
                )
                opening_count = (
                    len(schedule[date]["opening"])
                    if schedule[date]["opening"] is not None
                    else 0
                )
                middle_count = (
                    len(schedule[date]["middle"])
                    if schedule[date]["middle"] is not None
                    else 0
                )
                closing_count = (
                    len(schedule[date]["closing"])
                    if schedule[date]["closing"] is not None
                    else 0
                )

                ws.cell(
                    row=sum_row,
                    column=col_idx,
                    value=morning_count + opening_count + middle_count + closing_count,
                )

                ws.cell(row=sum_row + 1, column=1, value="MORNING")
                ws.cell(row=sum_row + 1, column=col_idx, value=morning_count)

                ws.cell(row=sum_row + 2, column=1, value="OPENING")
                ws.cell(row=sum_row + 2, column=col_idx, value=opening_count)

                ws.cell(row=sum_row + 3, column=1, value="MIDDAY")
                ws.cell(row=sum_row + 3, column=col_idx, value=middle_count)

                ws.cell(row=sum_row + 4, column=1, value="CLOSING")
                ws.cell(row=sum_row + 4, column=col_idx, value=closing_count)

        self.apply_excel_formatting(ws, all_dates, all_members)

        save_path = (
            f"{self.USERPATH}{self.MONTH_NAME.lower()}_schedule_{self.YEAR}.xlsx"
        )
        for date in schedule:
            if not self.is_weekend(date):
                for shift_type, staff_list in schedule[date].items():
                    if staff_list is not None:
                        for staff in staff_list:
                            if staff not in self.no_reply_members:
                                staff_shifts = staff_availability.get(staff, [])
                                available_shifts = [
                                    s
                                    for d, shifts in staff_shifts
                                    if d == date
                                    for s in shifts
                                ]
                                if shift_type not in available_shifts:
                                    print(
                                        f"Warning: {staff} assigned to {shift_type} shift on {date} but didn't sign up for it!"
                                    )

            save_path = (
                f"{self.USERPATH}{self.MONTH_NAME.lower()}_schedule_{self.YEAR}.xlsx"
            )
            print(f"\nSaving schedule to: {save_path}")
            wb.save(save_path)
        wb.save(save_path)


def main():
    scheduler = BarScheduler()
    scheduler.create_schedule()


if __name__ == "__main__":
    main()
